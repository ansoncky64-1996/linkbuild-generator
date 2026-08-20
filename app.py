"""
Digital Zoo Linkbuild Generator — Web UI
=========================================
streamlit run app.py
"""

import io
import json
import os
import time
import zipfile
import tempfile
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

import streamlit as st

st.set_page_config(
    page_title="DZ Linkbuild Generator",
    page_icon="🔗",
    layout="wide",
)

from generate import (
    PLACEHOLDER_URL,
    build_docx_file,
    build_single_docx,
    detect_language,
    generate_article_content,
    parse_excel,
)

# ================================================================
# Secrets
# ================================================================
try:
    API_KEY = st.secrets["OPENROUTER_API_KEY"]
except Exception:
    API_KEY = os.environ.get("OPENROUTER_API_KEY", "")

if not API_KEY:
    st.error("⚠️ 請設定 OPENROUTER_API_KEY（Streamlit Secrets 或環境變數）")
    st.stop()

DEFAULT_MODEL = "~deepseek/deepseek-v4-flash-latest"
try:
    CONFIGURED_MODEL = st.secrets.get("LB_MODEL", DEFAULT_MODEL)
except Exception:
    CONFIGURED_MODEL = os.environ.get("LB_MODEL", DEFAULT_MODEL)

# 每篇約 5k prompt token + 4k completion token（已計入平均 2.5 次 call）
# 成本 = 每月 4 個 batch × 20 篇 = 80 篇
MODEL_CHOICES = {
    "qwen/qwen3.8-max":
        "中文最強（阿里），指令跟得貼，約 $2.72/月　★ 建議",
    "deepseek/deepseek-v4-pro-0813":
        "同系列升級版，約 $1.62/月",
    "qwen/qwen3.8-27b":
        "細啲嘅 Qwen，抵用，約 $1.20/月",
    "bytedance-seed/seed-2-1-turbo":
        "中文原生（字節），約 $1.00/月",
    "x-ai/grok-4.6":
        "非中國出品，用詞較少大陸味，約 $2.72/月",
    "~deepseek/deepseek-v4-flash-latest":
        "最平但最唔聽話，約 $0.07/月",
}


# ================================================================
# Cached Excel parsing (only runs ONCE per file, not on every click)
# ================================================================
@st.cache_data(show_spinner="📊 讀取 Excel 中...")
def load_all_batches(file_bytes, sheet_name=None):
    """Parse Excel once and cache results. Re-runs only when file changes."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        batch_counts = {}
        warnings = []
        for b in range(1, 10):
            report = []
            arts = parse_excel(tmp_path, b, sheet_name, report=report)
            if arts:
                batch_counts[b] = arts
                warnings.extend(f"Batch {b}：{w}" for w in report)
        return batch_counts, warnings
    finally:
        os.unlink(tmp_path)


@st.cache_data(show_spinner=False)
def list_sheets(file_bytes):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        from openpyxl import load_workbook
        wb = load_workbook(tmp_path, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    finally:
        os.unlink(tmp_path)


# ================================================================
# UI
# ================================================================
st.title("🔗 DZ Linkbuild Generator")
st.caption("上傳 Excel → 選擇 Batch → 合規檢查 → 生成 .docx → 下載後拖入 Google Drive")
with st.expander("🤖 Model 設定", expanded=False):
    options = list(MODEL_CHOICES)
    if CONFIGURED_MODEL not in options:
        options.insert(0, CONFIGURED_MODEL)
    options.append("✏️ 自訂")
    picked = st.selectbox(
        "生成用嘅 model",
        options,
        index=options.index(CONFIGURED_MODEL),
        format_func=lambda m: (
            m if m == "✏️ 自訂" else f"{m}　—　{MODEL_CHOICES.get(m, 'Secrets 設定')}"
        ),
        help="改咗只影響今次 session。想長期改就去 Streamlit Secrets 設 LB_MODEL。",
    )
    MODEL = (
        st.text_input("自訂 model id", value=CONFIGURED_MODEL).strip()
        if picked == "✏️ 自訂" else picked
    )
    st.caption(
        "幾乎所有新 model 都係 reasoning model，本工具會送 "
        "`reasoning={\"enabled\": false}` 同較大嘅 `max_tokens`，"
        "provider 唔收就自動除返個參數重試。"
    )
    st.caption(
        "⚠️ 大陸出品嘅 model 會寫「視頻／信息／質量／帶寬」等大陸用語。"
        "OpenCC 只轉字形唔轉用詞，所以本工具會另外自動換返香港用語，"
        "有歧義嘅（設置／水平）會喺合規警告度提示。"
    )

st.caption(f"🤖 而家用緊：`{MODEL}`")

# Streamlit Cloud 一斷線就會終止 script run,所有 local 變數即刻冇。
# 所以每篇一生成好就即刻寫入 session_state,download 掣亦由 session_state
# 渲染 —— 就算行到一半斷咗,已完成嗰啲仍然攞得返。
st.session_state.setdefault("done", {})     # 合規稿 {(batch, num): (article, content)}
st.session_state.setdefault("drafts", {})   # 未合規草稿,同樣要保住
st.session_state.setdefault("logs", {})     # {(batch, num): [log lines]}

# ── Upload ──
excel_file = st.file_uploader("📊 上傳 Linkbuilding Excel", type=["xlsx"])

if not excel_file:
    st.info("👆 請先上傳 Excel 檔案")
    st.stop()

file_bytes = excel_file.getvalue()
sheet_names = list_sheets(file_bytes)
sheet_name = st.selectbox("📄 選擇 Sheet", sheet_names, index=0)

batch_counts, parse_warnings = load_all_batches(file_bytes, sheet_name)

if parse_warnings:
    with st.expander(f"⚠️ Excel 讀取警告（{len(parse_warnings)} 項）", expanded=True):
        for w in parse_warnings:
            st.warning(w)

if not batch_counts:
    st.error(f"Sheet「{sheet_name}」搵唔到任何 Batch 資料。請確認 A 欄有「Batch N」標籤。")
    st.stop()

# ── Select (instant, no re-parsing) ──
st.divider()

batch_options = list(batch_counts.keys())
batch_labels = [
    f"Batch {b} — {len(arts)} 篇（#{arts[0]['number']}-#{arts[-1]['number']}）"
    for b, arts in batch_counts.items()
]
selected_idx = st.selectbox(
    "📁 選擇 Batch",
    range(len(batch_options)),
    format_func=lambda i: batch_labels[i],
)
selected_batch = batch_options[selected_idx]
articles = batch_counts[selected_batch]

first_num = articles[0]["number"]
last_num = articles[-1]["number"]

select_mode = st.radio(
    "選擇方式",
    ["📏 連續範圍", "🔢 指定文章編號"],
    horizontal=True,
)

if select_mode == "📏 連續範圍":
    col_s, col_e = st.columns(2)
    with col_s:
        start_num = st.number_input(
            "起始 #", min_value=first_num, max_value=last_num, value=first_num,
        )
    with col_e:
        end_num = st.number_input(
            "結束 #", min_value=first_num, max_value=last_num, value=last_num,
        )
    filtered = [a for a in articles if start_num <= a["number"] <= end_num]
else:
    article_nums_input = st.text_input(
        "輸入文章編號（用逗號分隔）",
        placeholder=f"例如：{first_num}, {first_num+2}, {first_num+5}",
    )
    if article_nums_input:
        try:
            selected_nums = set(
                int(n.strip()) for n in article_nums_input.split(",") if n.strip()
            )
            filtered = [a for a in articles if a["number"] in selected_nums]
            not_found = selected_nums - {a["number"] for a in filtered}
            if not_found:
                st.warning(f"以下編號在此 Batch 中找不到：{sorted(not_found)}")
        except ValueError:
            st.error("格式錯誤，請輸入數字並用逗號分隔")
            filtered = []
    else:
        filtered = []

# ── Preview ──
with st.expander(f"📋 預覽（{len(filtered)} 篇）", expanded=False):
    preview_data = []
    for a in filtered:
        lang = detect_language(a["keyword1"] + a.get("keyword2", ""))
        preview_data.append({
            "#": a["number"],
            "Keyword 1": a["keyword1"],
            "Keyword 2": a["keyword2"] or "—",
            "Category": a["category"],
            "語言": "中文" if lang == "zh-HK" else "EN",
            "備註": "；".join(a.get("warnings", [])) or "—",
        })
    st.dataframe(preview_data, width="stretch", hide_index=True)

# ── Excel-level 警告 ──
placeholder_articles = [
    a["number"] for a in filtered
    if PLACEHOLDER_URL in (a["url1"], a["url2"])
]
if placeholder_articles:
    st.warning(
        f"⚠️ 以下文章喺 Excel 冇 target URL，會用 placeholder "
        f"`{PLACEHOLDER_URL}`，交稿前必須換返真實連結：{placeholder_articles}"
    )

# ── Settings ──
with st.expander("⚙️ 進階設定", expanded=False):
    parallel = st.slider(
        "同時生成篇數", 1, 5, 3,
        help="同時呼叫 API 的數量。越大越快，但太高可能觸發 rate limit。",
    )
    stagger = st.slider(
        "每次呼叫間隔（秒）", 0.0, 5.0, 1.5, step=0.5,
        help="避免同時打爆 OpenRouter rate limit。",
    )
    FAIL_STREAK_LIMIT = st.slider(
        "連續失敗幾多篇就自動停", 2, 20, 5,
        help="個 model 唔啱用嘅時候，唔好燒足幾個鐘先知。",
    )
    skip_done = st.checkbox(
        "跳過今個 session 已經生成好嘅文章", value=True,
        help="斷咗線之後再撳一次，就只會補做未完成嗰啲。",
    )

# ── Buttons ──
st.divider()
no_articles = len(filtered) == 0
col_a, col_b, col_spacer = st.columns([1, 1, 2])
with col_a:
    btn_generate = st.button("🚀 生成文章", type="primary", width="stretch", disabled=no_articles)
with col_b:
    btn_dry = st.button("📝 Dry Run（預覽文字）", width="stretch", disabled=no_articles)

if no_articles and select_mode == "🔢 指定文章編號":
    st.info("👆 請輸入要生成的文章編號")


# ================================================================
# Generation
# ================================================================
def _summarise(log):
    """由 log 抽出一句人睇得明嘅失敗原因。"""
    text = "\n".join(log)
    if "429" in text or "rate" in text.lower():
        return "API rate limit（調低同時生成篇數 / 加大間隔）"
    if "OpenRouter 4" in text or "OpenRouter 5" in text:
        line = next((l for l in log if "OpenRouter" in l), "")
        return f"API 出錯 — {line.split('OpenRouter', 1)[-1].strip()[:120]}"
    if "API 出錯" in text or "Timeout" in text or "ConnectionError" in text:
        return "API 連線失敗 / timeout"
    if "JSONDecodeError" in text or "Expecting" in text:
        return "Model 回覆唔係合法 JSON"
    bullets = [l.strip("• ").strip() for l in log if l.strip().startswith("•")]
    if bullets:
        top = {}
        for b in bullets:
            key = ("字數" if "字數" in b else
                   "關鍵字字面重複" if "字面出現" in b else
                   "marker 缺失" if "搵唔到 marker" in b else
                   "關鍵字入咗標題" if "H1 或 H2" in b else
                   "H1 有標點" if "H1 唔准有標點" in b else
                   "H1 過長" if "H1 長度" in b else
                   "簡體字" if "簡體" in b else
                   "廣東話口語" if "廣東話" in b else
                   "英文稿夾中文" if "中文字" in b else
                   "破折號" if "破折號" in b else
                   "Keyword buffer" if "buffer" in b.lower() else b[:40])
            top[key] = top.get(key, 0) + 1
        worst = max(top.items(), key=lambda kv: kv[1])[0]
        return f"合規過唔到 — 主要係「{worst}」"
    return "生成失敗（詳情見下面 log）"



# ================================================================
# 交付檔案（由 session_state 渲染，斷線都唔會冇）
# ================================================================
@st.cache_data(show_spinner=False)
def _build_bytes_cached(payload_json):
    """Streamlit 每次互動都會重跑成個 script，冇 cache 就會不停重 build。
    payload_json 係 hashable，內容一樣就直接攞返 cache。"""
    return _build_bytes(json.loads(payload_json))


def _build_bytes(pairs):
    """回 (合併 docx bytes, 每篇一個檔嘅 zip bytes)。"""
    with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
        build_docx_file(pairs, tmp.name)
        combined_path = tmp.name
    with open(combined_path, "rb") as f:
        combined = f.read()
    os.unlink(combined_path)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for article, content in pairs:
            with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as t:
                build_single_docx(article, content, t.name)
                single = t.name
            with open(single, "rb") as f:
                zf.writestr(
                    f"{article['number']:03d}_{article['keyword1'][:20]}.docx", f.read()
                )
            os.unlink(single)
    return combined, buf.getvalue()


def render_drafts(batch):
    """未合規草稿。獨立一區、獨立檔案，永遠唔會混入交付稿。"""
    pairs = sorted(
        ((a, c) for (b, _n), (a, c) in st.session_state["drafts"].items() if b == batch),
        key=lambda x: x[0]["number"],
    )
    if not pairs:
        return
    st.divider()
    st.subheader(f"⚠️ 未合規草稿（{len(pairs)} 篇，要人手執）")
    st.caption("呢啲文生成到，但過唔到合規檢查。**唔會混入上面份交付稿**。")

    buckets = {}
    for a, c in pairs:
        for f in c.get("_fails", []):
            key = ("字數超標" if "超出" in f else
                   "字數不足" if "唔夠" in f else
                   "marker 缺失" if "搵唔到 marker" in f else
                   "關鍵字字面重複" if "字面出現" in f else
                   "keyword buffer" if "buffer" in f.lower() else f[:30])
            buckets.setdefault(key, []).append(a["number"])
    for reason, nums in sorted(buckets.items(), key=lambda kv: -len(kv[1])):
        st.warning(f"**{reason}** — {len(nums)} 篇：{sorted(set(nums))}")

    with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as t:
        build_docx_file(pairs, t.name)
        path = t.name
    with open(path, "rb") as f:
        data = f.read()
    os.unlink(path)
    st.download_button(
        f"⬇️ 下載未合規草稿（{len(pairs)} 篇）", data=data,
        file_name=f"Batch_{batch}_未合規草稿.docx",
        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


def render_deliverables(batch):
    """由 session_state 攞返呢個 batch 已完成嘅文章，永遠提供下載。"""
    pairs = sorted(
        ((a, c) for (b, _n), (a, c) in st.session_state["done"].items() if b == batch),
        key=lambda x: x[0]["number"],
    )
    if not pairs:
        return

    st.divider()
    nums = [a["number"] for a, _ in pairs]
    st.subheader(f"📦 已生成（{len(pairs)} 篇）")
    st.caption(f"文章編號：{nums}　—　呢個清單存喺 session，斷線／重跑都仲喺度。")

    with st.spinner("📄 建立 Word 文件中..."):
        combined, zipped = _build_bytes_cached(json.dumps(pairs, ensure_ascii=False))

    name = f"Combined_Batch_{batch}_#{nums[0]}-{nums[-1]}"
    c1, c2, c3 = st.columns([2, 2, 1])
    with c1:
        st.download_button(
            f"⬇️ {name}.docx（合併稿）", data=combined, file_name=f"{name}.docx",
            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            type="primary", width="stretch",
        )
        st.caption("💡 下載後拖入 Google Drive → 自動轉為 Google Doc")
    with c2:
        st.download_button(
            f"⬇️ {name}_single.zip（每篇一個檔）", data=zipped,
            file_name=f"{name}_single.zip", mime="application/zip", width="stretch",
        )
        st.caption("💡 呢個 zip 可以直接跑 `validate.py`")
    with c3:
        if st.button("🗑️ 清空", width="stretch"):
            for k in [k for k in st.session_state["done"] if k[0] == batch]:
                del st.session_state["done"][k]
            st.rerun()

    placeholders = [
        a["number"] for a, _ in pairs if PLACEHOLDER_URL in (a["url1"], a["url2"])
    ]
    if placeholders:
        st.warning(f"⚠️ 仲用緊 placeholder URL：{placeholders}")


_rate_lock = threading.Lock()
_last_call = [0.0]


def _throttle(min_gap):
    """全域節流，避免 parallel worker 同一刻打爆 rate limit。"""
    if min_gap <= 0:
        return
    with _rate_lock:
        wait = _last_call[0] + min_gap - time.monotonic()
        if wait > 0:
            time.sleep(wait)
        _last_call[0] = time.monotonic()


def _generate_one(article, min_gap):
    """Worker function for parallel generation。log 會一齊帶返去畀 UI 顯示。"""
    _throttle(min_gap)
    log = []
    content = generate_article_content(article, API_KEY, MODEL, log=log)
    return article, content, log


if btn_generate or btn_dry:
    is_dry_run = btn_dry
    if skip_done and not is_dry_run:
        already = {n for b, n in st.session_state["done"] if b == selected_batch}
        skipped = [a["number"] for a in filtered if a["number"] in already]
        if skipped:
            st.info(f"⏭️ 跳過已完成：{skipped}")
        filtered = [a for a in filtered if a["number"] not in already]
        total = len(filtered)
    articles_with_content = []
    noncompliant = []
    failed = []
    total = len(filtered)

    progress_bar = st.progress(0, text="準備中...")
    status_area = st.container()

    def _persist(article, content, log):
        """合規同未合規要分開存,否則未合規草稿會混入交付稿。"""
        key = (selected_batch, article["number"])
        st.session_state["logs"][key] = log
        if not content:
            return
        if content.get("_compliant", True):
            st.session_state["done"][key] = (article, content)
            st.session_state["drafts"].pop(key, None)
        else:
            st.session_state["drafts"][key] = (article, content)

    def _report(article, content, log):
        num = article["number"]
        _persist(article, content, log)
        if not content:
            failed.append((num, log))
            reason = _summarise(log)
            with status_area:
                st.error(f"❌ #{num} — {reason}")
                with st.expander(f"🔍 #{num} 失敗詳情（{len(log)} 行）", expanded=False):
                    st.code("\n".join(log) or "（冇任何訊息）", language="text")
            return
        if not content.get("_compliant", True):
            noncompliant.append((article, content))
            with status_area:
                st.warning(
                    f"⚠️ #{num} — {content['h1']}　"
                    f"（{content['_word_count']} 字，**{len(content.get('_fails', []))} 項未過**，要人手執）"
                )
                for f in content.get("_fails", []):
                    st.caption(f"　　• {f}")
                with st.expander(f"🔍 #{num} 生成過程（{len(content.get('_log', []))} 行）"):
                    st.code("\n".join(content.get("_log", [])), language="text")
            return
        articles_with_content.append((article, content))
        unit = "words" if content["_lang"] == "en" else "字"
        with status_area:
            st.success(
                f"✅ #{num} — {content['h1']}　"
                f"（{content['_word_count']} {unit}，合規通過）"
            )
            for w in content.get("_warnings", []):
                st.caption(f"　　⚠️ {w}")
            if content.get("_log"):
                with st.expander(f"🔍 #{num} 生成過程（{len(content['_log'])} 行）", expanded=False):
                    st.code("\n".join(content["_log"]), language="text")

    consecutive_fail = [0]
    aborted = [False]

    def _breaker():
        """連續失敗太多就即刻收手，唔好燒足三個鐘先發現個 model 唔work。"""
        if consecutive_fail[0] >= FAIL_STREAK_LIMIT:
            aborted[0] = True
            return True
        return False

    if parallel <= 1:
        for i, article in enumerate(filtered):
            if _breaker():
                break
            label = f"#{article['number']}: {article['keyword1']}"
            progress_bar.progress(i / total, text=f"⏳（{i+1}/{total}）{label}")
            art, content, log = _generate_one(article, stagger)
            if content and content.get("_compliant", True):
                consecutive_fail[0] = 0
            else:
                consecutive_fail[0] += 1
            _report(art, content, log)
    else:
        done_count = 0
        with ThreadPoolExecutor(max_workers=parallel) as executor:
            futures = [
                executor.submit(_generate_one, art, stagger)
                for art in filtered
            ]
            for future in as_completed(futures):
                article, content, log = future.result()
                if content and content.get("_compliant", True):
                    consecutive_fail[0] = 0
                else:
                    consecutive_fail[0] += 1
                done_count += 1
                if _breaker():
                    for f in futures:
                        f.cancel()
                progress_bar.progress(
                    done_count / total,
                    text=f"⏳（{done_count}/{total}）已完成 #{article['number']}",
                )
                _report(article, content, log)

        articles_with_content.sort(key=lambda x: x[0]["number"])

    if aborted[0]:
        st.error(
            f"🛑 連續 {FAIL_STREAK_LIMIT} 篇失敗，已經自動停低，唔好再燒 token。"
            "　請睇下面嘅失敗原因，換個 model 或者調整設定再試。"
            "　**已完成嘅文章唔會冇咗**，喺下面「📦 已生成」度攞。"
        )

    progress_bar.progress(
        1.0,
        text=f"✅ 合規 {len(articles_with_content)}/{total} 篇"
             + (f"，未合規 {len(noncompliant)} 篇" if noncompliant else "")
             + (f"，失敗 {len(failed)} 篇" if failed else ""),
    )

    # ── Output ──
    st.divider()

    if is_dry_run:
        if not articles_with_content:
            st.error("所有文章都生成失敗")
        else:
            st.subheader("📝 文字預覽")
            for article, content in articles_with_content:
                with st.expander(f"#{article['number']} — {content['h1']}", expanded=False):
                    text = f"[H1: {content['h1']}]\n\n"
                    for sec in content["sections"]:
                        if sec.get("h2"):
                            text += f"[H2: {sec['h2']}]\n\n"
                        text += sec.get("body", "") + "\n\n"
                    st.text(text)
    elif articles_with_content:
        st.balloons()
        st.success(f"🎉 今次生成咗 {len(articles_with_content)} 篇，全部通過合規檢查")
    elif not noncompliant and not failed:
        st.info("今次冇生成任何新文章")

    if failed:
        st.divider()
        st.subheader(f"❌ 完全失敗（{len(failed)} 篇）")
        buckets = {}
        for num, log in failed:
            buckets.setdefault(_summarise(log), []).append(num)
        for reason, nums in sorted(buckets.items(), key=lambda kv: -len(kv[1])):
            st.warning(f"**{reason}** — {len(nums)} 篇：{nums}")
        st.download_button(
            "⬇️ 下載完整失敗 log（.txt）",
            data="\n\n".join(
                f"===== #{num} =====\n" + "\n".join(log) for num, log in failed
            ),
            file_name="linkbuild_failures.log",
            mime="text/plain",
        )


# ── 永遠渲染（就算今次冇按生成，之前做好嘅都攞得返）──
if not btn_dry:
    render_deliverables(selected_batch)
    render_drafts(selected_batch)
